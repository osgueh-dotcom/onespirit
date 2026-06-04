import openpyxl
from datetime import datetime, date
from decimal import Decimal
import sys
import os

# Set sys.path to backend directory to import app modules correctly
backend_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.append(backend_path)
sys.stdout.reconfigure(encoding='utf-8')

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from app.core.config import settings
from app.core.database import Base
from app.modules.auth.models import User, Role
from app.modules.crm.models import Customer, Contact
from app.modules.projects.models import Project, ProjectStatusLog
from app.modules.events.models import EventSchedule
from app.modules.tasks.models import Task
from app.modules.finance.models import Invoice, Payment
from app.modules.documents.models import Document

# 1. Connect to Database (Smart fallback between 'db', 'localhost' PostgreSQL, and SQLite)
db_url = settings.DATABASE_URL
engine = None

try:
    if "db" in db_url:
        try:
            engine = create_engine(db_url, connect_args={"connect_timeout": 2})
            engine.connect()
            print("Connected directly to docker host 'db' PostgreSQL database.")
        except Exception:
            fallback_url = db_url.replace("@db:", "@localhost:")
            engine = create_engine(fallback_url, connect_args={"connect_timeout": 2})
            engine.connect()
            db_url = fallback_url
            print("Docker host 'db' unavailable. Connected to localhost PostgreSQL database.")
except Exception:
    pass

if not engine:
    db_url = "sqlite:///./onespirit_sqlite.db"
    print("PostgreSQL database unavailable. Falling back to local SQLite database: onespirit_sqlite.db")
    engine = create_engine(db_url, connect_args={"check_same_thread": False})

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
db = SessionLocal()

# 2. Setup Base Admin User
# Ensure tables are built
Base.metadata.create_all(bind=engine)

# Look up admin
admin_user = db.query(User).filter(User.email == "admin@onespirit.asia").first()
if not admin_user:
    # Seed standard admin role and admin user if they don't exist
    from app.modules.auth.service import seed_roles_and_admin
    seed_roles_and_admin(db)
    admin_user = db.query(User).filter(User.email == "admin@onespirit.asia").first()

print(f"Using default PM/Admin account: {admin_user.full_name}")

# 3. Open Excel file
excel_path = r"d:\One Spirit\Project Workflow OS 2025.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb["Workflow"]

# Fetch row 5 for header indexing
row5 = list(sheet.iter_rows(min_row=5, max_row=5, values_only=True))[0]
headers = {str(col).strip().replace(chr(10), ' '): idx for idx, col in enumerate(row5) if col is not None}

def get_val(r, header_name, default=None):
    idx = headers.get(header_name)
    if idx is not None and idx < len(r):
        return r[idx]
    return default

def parse_decimal(val):
    if val is None:
        return Decimal("0.00")
    try:
        cleaned = str(val).replace("Rp", "").replace(",", "").replace(" ", "").strip()
        if not cleaned or cleaned.lower() in ["n/a", "-", "free", "tbd", "none"]:
            return Decimal("0.00")
        return Decimal(cleaned)
    except Exception:
        return Decimal("0.00")

print("\n--- Starting Data Import ---")
rows = list(sheet.iter_rows(min_row=6, values_only=True))

customer_count = 0
project_count = 0
event_count = 0
invoice_count = 0
doc_count = 0

for i, r in enumerate(rows):
    # Retrieve core customer details
    company_name = get_val(r, 'Company Name')
    if not company_name:
        continue  # Skip rows without company name
        
    company_name = str(company_name).strip()
    
    category = get_val(r, 'Customer Category', 'Corporate')
    category = str(category or 'Corporate').strip()
    
    cp_name = get_val(r, 'CP')
    cp_phone = get_val(r, 'Phone Number')
    cp_email = get_val(r, 'Email')
    
    # Check/Create Customer
    customer = db.query(Customer).filter(Customer.company_name == company_name, Customer.deleted_at == None).first()
    if not customer:
        customer = Customer(
            company_name=company_name,
            category=category,
            notes="Imported from 2025 Excel workflow"
        )
        db.add(customer)
        db.commit()
        db.refresh(customer)
        customer_count += 1
        print(f"Registered Customer: {company_name}")
        
    # Check/Create Contact Person
    if cp_name:
        cp_name = str(cp_name).strip()
        contact = db.query(Contact).filter(
            Contact.customer_id == customer.id,
            Contact.name == cp_name,
            Contact.deleted_at == None
        ).first()
        if not contact:
            contact = Contact(
                customer_id=customer.id,
                name=cp_name,
                phone=str(cp_phone or '').strip() or None,
                email=str(cp_email or '').strip() or None,
                position="Point of Contact"
            )
            db.add(contact)
            db.commit()
            
    # Retrieve project details
    event_category = get_val(r, 'Event Category', 'Event')
    event_category = str(event_category or 'Operational Program').strip()
    
    prog = get_val(r, 'Prog', '')
    prog_str = f" ({prog})" if prog else ""
    
    proj_title = f"{company_name} - {event_category}{prog_str}"
    
    # Map project status
    raw_status = str(get_val(r, 'Project Status', 'inquiry')).strip().lower()
    status_mapping = {
        'inquery': 'inquiry',
        'inquiry': 'inquiry',
        'quotation': 'quotation',
        'negotiation': 'negotiation',
        'confirmed': 'confirmed',
        'ongoing': 'ongoing',
        'completed': 'completed',
        'canceled': 'canceled',
        'lost': 'canceled'
    }
    status = status_mapping.get(raw_status, 'inquiry')
    
    # Map quotation status
    raw_q_status = str(get_val(r, 'Quotation Status', 'draft')).strip().lower()
    q_status_mapping = {
        'approved': 'approved',
        'sent': 'sent',
        'draft': 'draft',
        'rejected': 'rejected'
    }
    q_status = q_status_mapping.get(raw_q_status, 'draft')
    
    # Map dates
    raw_start = get_val(r, 'Event Date Start')
    raw_end = get_val(r, 'Event Date End')
    
    start_date = None
    if isinstance(raw_start, datetime):
        start_date = raw_start.date()
    elif isinstance(raw_start, date):
        start_date = raw_start
        
    end_date = None
    if isinstance(raw_end, datetime):
        end_date = raw_end.date()
    elif isinstance(raw_end, date):
        end_date = raw_end

    # Map budget
    raw_budget = get_val(r, 'Budget (Rp)', 0)
    budget = parse_decimal(raw_budget)
    
    # Check/Create Project
    project = db.query(Project).filter(
        Project.customer_id == customer.id,
        Project.title == proj_title,
        Project.deleted_at == None
    ).first()
    
    if not project:
        project = Project(
            title=proj_title,
            status=status,
            quotation_status=q_status,
            start_date=start_date,
            end_date=end_date,
            budget=budget,
            revenue=budget,
            customer_id=customer.id,
            created_by_id=admin_user.id,
            assigned_to_id=admin_user.id
        )
        db.add(project)
        db.commit()
        db.refresh(project)
        project_count += 1
        print(f"Created Project: '{proj_title}' [Status: {status}]")
        
        # Log initial transition
        log = ProjectStatusLog(
            project_id=project.id,
            from_status="none",
            to_status=status,
            notes="Imported from 2025 Excel workflow",
            changed_by_id=admin_user.id
        )
        db.add(log)
        db.commit()

    # Link venue & schedules if venue is set
    venue = get_val(r, 'Venue')
    if venue and start_date:
        venue_str = str(venue).strip()
        existing_schedule = db.query(EventSchedule).filter(
            EventSchedule.project_id == project.id,
            EventSchedule.venue_name == venue_str,
            EventSchedule.deleted_at == None
        ).first()
        if not existing_schedule:
            # Create interactive schedule
            start_dt = datetime.combine(start_date, datetime.min.time())
            end_dt = datetime.combine(end_date or start_date, datetime.max.time())
            schedule = EventSchedule(
                project_id=project.id,
                venue_name=venue_str,
                start_time=start_dt,
                end_time=end_dt,
                pic_id=admin_user.id,
                rundown=[
                    {"time": "08:00", "activity": "Gathering & Briefing", "pic": "Lead", "notes": "Initial setup"},
                    {"time": "10:00", "activity": "Core Program Activities", "pic": "Lead", "notes": "Main operational items"}
                ]
            )
            db.add(schedule)
            db.commit()
            event_count += 1
            
    # Link Quotation/Invoice details
    q_num = get_val(r, 'Quotation Number')
    if q_num:
        q_num_str = str(q_num).strip()
        invoice = db.query(Invoice).filter(Invoice.invoice_number == q_num_str, Invoice.deleted_at == None).first()
        if not invoice:
            # Auto issue invoice
            invoice = Invoice(
                project_id=project.id,
                invoice_number=q_num_str,
                amount=budget,
                issue_date=start_date or date.today(),
                due_date=start_date or date.today(),
                status="unpaid",
                notes="Auto-generated from Excel import"
            )
            db.add(invoice)
            db.commit()
            invoice_count += 1

    # Link Google Drive links / references
    link_photo = get_val(r, 'Link Photo')
    link_video = get_val(r, 'Link Video')
    link_teaser = get_val(r, 'Link Teaser')
    
    links = [
        ('Link Photo', link_photo, 'image'),
        ('Link Video', link_video, 'video'),
        ('Link Teaser', link_teaser, 'teaser')
    ]
    for label, val, f_type in links:
        if val and str(val).startswith("http"):
            val_str = str(val).strip()
            existing_doc = db.query(Document).filter(
                Document.project_id == project.id,
                Document.file_path == val_str,
                Document.deleted_at == None
            ).first()
            if not existing_doc:
                doc = Document(
                    project_id=project.id,
                    title=f"Archived {label}",
                    file_path=val_str,
                    file_type=f_type,
                    storage_type="google_drive",
                    uploaded_by_id=admin_user.id,
                    notes=f"Archived {label} link imported from Excel"
                )
                db.add(doc)
                db.commit()
                doc_count += 1

db.close()
print("\n--- Data Import Completed ---")
print(f"Registered Customers: {customer_count}")
print(f"Created Projects: {project_count}")
print(f"Linked Venue Schedules: {event_count}")
print(f"Issued Invoices: {invoice_count}")
print(f"Archived Documents: {doc_count}")
