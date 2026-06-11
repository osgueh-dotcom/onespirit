import io
import uuid
import openpyxl
from datetime import datetime, date, timezone
from typing import List, Dict, Any, Tuple, Optional
from sqlalchemy.orm import Session

from app.modules.auth.models import User
from app.modules.crm.models import Customer, Contact
from app.modules.projects.models import Project, ProjectStatusLog, ProjectInstrument
from app.modules.event_sources.models import EventSource
from app.modules.events.models import EventSchedule
from app.modules.tasks.models import Task
from app.modules.finance.models import Invoice, Payment
from app.modules.documents.models import Document
from app.core.activity import log_activity
from app.modules.projects.service import check_project_validation_warnings
from app.modules.crm.service import check_duplicate_customer, normalize_customer_name

class TempProject(Project):
    def __init__(self, *args, **kwargs):
        self._simulated_paid = kwargs.pop('simulated_paid', 0.0)
        super().__init__(*args, **kwargs)
    
    @property
    def paid_amount(self) -> float:
        return self._simulated_paid

# Enriched Mapping lists for column headers (case-insensitive & robust)
HEADER_MAP = {
    "company_name": ["company name", "company", "client", "customer name", "customer"],
    "customer_category": ["customer category", "client category", "category", "classification"],
    "contact_name": ["contact person", "contact name", "contact", "pic customer", "cp"],
    "phone": ["phone number", "phone", "contact phone", "mobile"],
    "email": ["email", "email address", "contact email"],
    
    # Refactored mappings
    "project_code": ["project code", "code", "proj code", "project_code"],
    "inquiry_date": ["inq date", "inquiry date", "inq_date", "inquiry", "inq. date"],
    "partner": ["partner", "vendor", "source", "event source", "event_source"],
    "sales": ["sales", "sales name", "external sales", "sales_name", "external sales name"],
    "po": ["po", "program owner", "program_owner", "owner"],
    "pm": ["pm", "program manager", "program_manager", "manager"],
    "event_category": ["event category", "program category", "event_category"],
    "program_type": ["program type", "type", "program_type"],
    "program_name": ["program name", "project name", "title", "program", "project title", "prog", "prog name"],
    "quantity": ["quantity", "qty", "pax", "total pax", "jumlah pax"],
    "venue": ["venue", "location", "place", "event venue"],
    "duration": ["duration", "days", "durasi"],
    "event_date_start": ["start date", "event date", "start", "proposed start date", "event date start", "date start"],
    "event_date_end": ["end date", "end", "proposed end date", "event date end", "date end"],
    "start_date": ["start date", "event date", "start", "proposed start date", "event date start", "date start"],
    "end_date": ["end date", "end", "proposed end date", "event date end", "date end"],
    "quotation_date": ["quotation date", "quotation_date", "quote date"],
    "quotation_number": ["quotation number", "quotation no", "quote no", "quotation reference", "quote number"],
    "quotation_status": ["quotation status", "quote status", "quotation_status"],
    "program_status": ["program status", "status", "program_status", "workflow status"],
    "payment_status": ["payment status", "pay status", "payment_status"],
    "project_status": ["project status", "project_status"],
    "budget": ["budget", "value", "deal value", "deal_value", "price", "budget (rp)", "budget rp"],
    
    "cl": ["cl", "cl checklist"],
    "ros": ["ros", "ros checklist"],
    "ck": ["ck", "ck checklist"],
    "pnl": ["pnl", "pnl checklist", "pnl sheet"],
    "pf": ["pf", "pf checklist"],
    "matrix": ["matrix", "matrix checklist"],
    "photo_links": ["photo links", "photos", "photo link", "link photo"],
    "video_links": ["video links", "videos", "video link", "link video"],
    "teaser_links": ["teaser links", "teaser", "teaser link", "link teaser"],
    "ig_links": ["ig", "instagram", "ig link", "instagram link", "ig_links"],
    "yt_links": ["yt", "youtube", "yt link", "youtube link", "yt_links"],
    "invoice_status": ["invoice status", "invoicing status", "billing status"],
    "payment_schedule": ["payment schedule", "payment schadule", "payment plan"],
    "signed_document": ["signed document", "signed file", "contract", "signed agreement"],
    "cancel_reason": ["cancel reason", "cancel_reason", "reason cancel"],
    "mom_notes": ["mom", "mom notes", "minutes of meeting", "mom_notes"],
    "general_notes": ["general notes", "notes", "general_notes"]
}

def clean_header(val: Any) -> str:
    if not val:
        return ""
    import re
    cleaned = str(val).strip().lower().replace("_", " ").replace("-", " ")
    return re.sub(r'\s+', ' ', cleaned)

def find_column_indices(sheet: openpyxl.worksheet.worksheet.Worksheet) -> Tuple[Dict[str, int], int]:
    best_row_idx = 1
    best_match_count = -1
    best_indices = {}

    for r_idx in range(1, min(15, sheet.max_row + 1)):
        row_cells = [cell.value for cell in sheet[r_idx]]
        current_indices = {}
        match_count = 0
        
        for key, aliases in HEADER_MAP.items():
            found = False
            for idx, val in enumerate(row_cells):
                cleaned = clean_header(val)
                if cleaned in aliases:
                    current_indices[key] = idx + 1  # 1-indexed for openpyxl
                    match_count += 1
                    found = True
                    break
            if not found:
                current_indices[key] = None
        
        if match_count > best_match_count:
            best_match_count = match_count
            best_row_idx = r_idx
            best_indices = current_indices

    return best_indices, best_row_idx

def get_row_val(row: Tuple, col_idx: int) -> Any:
    if not col_idx or col_idx > len(row):
        return None
    return row[col_idx - 1].value

def parse_date(val: Any) -> Any:
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val
    try:
        return datetime.strptime(str(val).split(" ")[0], "%Y-%m-%d").date()
    except Exception:
        try:
            return datetime.strptime(str(val).split(" ")[0], "%d/%m/%Y").date()
        except Exception:
            return None

def parse_float(val: Any) -> float:
    if not val:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        cleaned = str(val).replace("$", "").replace("Rp", "").replace(" ", "").strip()
        if not cleaned:
            return 0.0
        
        if "," in cleaned and "." in cleaned:
            if cleaned.find(",") < cleaned.find("."):
                cleaned = cleaned.replace(",", "")
            else:
                cleaned = cleaned.replace(".", "").replace(",", ".")
        elif "," in cleaned:
            parts = cleaned.split(",")
            if len(parts) == 2 and len(parts[1]) <= 2:
                cleaned = cleaned.replace(",", ".")
            else:
                cleaned = cleaned.replace(",", "")
        elif "." in cleaned:
            parts = cleaned.split(".")
            if len(parts) > 2:
                cleaned = cleaned.replace(".", "")
            elif len(parts) == 2 and len(parts[1]) > 2:
                cleaned = cleaned.replace(".", "")
                
        return float(cleaned)
    except Exception:
        return 0.0

# Mappings for controlled status taxonomies
def map_quotation_status(val: Any) -> str:
    cleaned = str(val or "").strip().lower()
    mapping = {
        "draft": "Draft", "sent": "Sent", "follow up": "Follow Up", "followup": "Follow Up",
        "revision": "Revision", "signed & deal": "Signed & Deal", "signed": "Signed & Deal",
        "deal": "Signed & Deal", "cancel": "Cancel", "canceled": "Cancel"
    }
    return mapping.get(cleaned, "Draft")

def map_program_status(val: Any) -> str:
    cleaned = str(val or "").strip().lower()
    mapping = {
        "inquiry": "Inquiry", "confirmed": "Confirmed", "preparation": "Preparation",
        "ready": "Ready", "running": "Running", "completed": "Completed",
        "reporting": "Reporting", "closed": "Closed", "cancel": "Cancel", "canceled": "Cancel"
    }
    return mapping.get(cleaned, "Inquiry")

def map_payment_status(val: Any) -> str:
    cleaned = str(val or "").strip().lower()
    mapping = {
        "not invoiced": "Not Invoiced", "invoiced": "Invoice Sent", "invoice sent": "Invoice Sent",
        "partial": "Partial Paid", "partial paid": "Partial Paid", "paid": "Paid",
        "outstanding": "Outstanding",
        "out standing": "Outstanding",
        "out-standing": "Outstanding",
        "outstanding payment": "Outstanding",
        "out standing payment": "Outstanding",
        "overdue": "Overdue"
    }
    return mapping.get(cleaned, "Not Invoiced")

def map_project_status(val: Any) -> str:
    cleaned = str(val or "").strip().lower()
    mapping = {
        "open": "Open", "active": "Active", "reporting": "Reporting",
        "closed": "Closed", "cancel": "Canceled", "canceled": "Canceled"
    }
    return mapping.get(cleaned, "Open")

def build_project_title(company_name: str, event_category: Any, program_name: Any) -> str:
    c_name = str(company_name or "").strip()
    e_cat = str(event_category or "").strip()
    p_name = str(program_name or "").strip()
    
    if c_name and c_name.lower() in p_name.lower():
        return p_name
        
    parts = []
    if c_name:
        parts.append(c_name)
        
    prog_str = ""
    if e_cat and p_name:
        prog_str = f"{e_cat} ({p_name})"
    elif e_cat:
        prog_str = e_cat
    elif p_name:
        prog_str = p_name
        
    if prog_str:
        parts.append(prog_str)
        
    title = " - ".join(parts)
    return title if title else "Unnamed Project"

# User resolver helper by name or initials
def resolve_user_by_name_or_initials(db: Session, val: Any) -> Optional[uuid.UUID]:
    if not val:
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
        
    # 1. Match by initial_code if available
    user_by_initial = db.query(User).filter(
        User.initial_code.ilike(val_str)
    ).filter(User.deleted_at == None).first()
    if user_by_initial:
        return user_by_initial.id

    # 2. Match by full_name or email
    user = db.query(User).filter(
        (User.full_name.ilike(val_str)) | (User.email.ilike(val_str))
    ).filter(User.deleted_at == None).first()
    if user:
        return user.id
        
    # 3. Match initials dynamically (e.g. "JD" -> "John Doe")
    users = db.query(User).filter(User.deleted_at == None).all()
    for u in users:
        parts = u.full_name.split()
        initials = "".join([part[0] for part in parts if part]).upper()
        if initials == val_str.upper():
            return u.id
            
    return None

# Event source resolver (auto-provisions if missing)
def resolve_or_create_event_source(db: Session, partner_val: Any, sales_val: Any) -> Tuple[Optional[uuid.UUID], Optional[str]]:
    if not partner_val and not sales_val:
        return None, None
    
    partner_str = str(partner_val).strip() if partner_val else ""
    sales_str = str(sales_val).strip() if sales_val else ""
    
    # Determine controlled source type
    combined_val = (partner_str + " " + sales_str).strip().lower()
    source_type = "Direct"
    warning = None
    
    if combined_val:
        if "hotel" in combined_val:
            source_type = "Hotel"
        elif "direct" in combined_val:
            source_type = "Direct"
        elif "repeater" in combined_val:
            source_type = "Repeater"
        elif "partner" in combined_val:
            source_type = "Partner"
        elif "instagram" in combined_val or "ig" in combined_val:
            source_type = "Instagram"
        elif "web" in combined_val or "website" in combined_val:
            source_type = "Web"
        elif "other" in combined_val:
            source_type = "Other"
        else:
            # Check for exact matches
            matched = False
            for ct in ["Hotel", "Direct", "Repeater", "Partner", "Instagram", "Web", "Other"]:
                if ct.lower() == combined_val:
                    source_type = ct
                    matched = True
                    break
            if not matched:
                source_type = "Other"
                warning = f"Unknown source type for '{partner_val or sales_val}', falling back to 'Other'."
    else:
        source_type = "Direct"
        
    query = db.query(EventSource).filter(EventSource.deleted_at == None)
    if partner_str:
        query = query.filter(EventSource.vendor_name.ilike(partner_str))
    if sales_str:
        query = query.filter(EventSource.sales_name.ilike(sales_str))
        
    source = query.first()
    if source:
        # Update existing source type if it's currently Other or Direct and we matched something more specific
        if source.source_type != source_type and source.source_type in ["Other", "Direct", None] and source_type != "Other":
            source.source_type = source_type
            db.commit()
        return source.id, warning
        
    source = EventSource(
        source_type=source_type,
        vendor_name=partner_str if partner_str else None,
        sales_name=sales_str if sales_str else None,
        contact="Auto Contact",
        notes="Created during Excel import sync",
        is_active=True
    )
    db.add(source)
    db.commit()
    db.refresh(source)
    return source.id, warning

def map_excel_instrument_val(val: Any) -> Tuple[str, Optional[str], Optional[str], Optional[str]]:
    val_str = str(val or "").strip()
    if not val_str or val_str == "-":
        return "Not Started", None, None, None
        
    val_lower = val_str.lower()
    if val_lower in ["done", "yes", "ada", "ok", "complete", "completed", "1", "x", "true", "check"]:
        return "Done", None, None, None
    elif val_str.startswith("http"):
        return "Done", val_str, None, None
    elif val_lower in ["n/a", "not required"]:
        return "Not Required", None, None, None
    elif val_lower in ["revision", "revise", "need revision"]:
        return "Need Revision", None, None, None
    else:
        warning_msg = f"Ambiguous instrument value '{val_str}'"
        return "In Progress", None, val_str, warning_msg

def parse_excel_sheet(db: Session, file_bytes: bytes) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    
    sheet = None
    for s_name in wb.sheetnames:
        if clean_header(s_name) == "workflow" or s_name == "Workflow":
            sheet = wb[s_name]
            break
    if not sheet:
        sheet = wb.active

    indices, header_row_idx = find_column_indices(sheet)
    
    warnings = []
    preview_rows = []
    
    new_customers = set()
    new_projects = set()
    conflicts_count = 0
    total_rows = 0
    valid_rows_count = 0
    invalid_rows_count = 0

    for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
        row = list(sheet[r_idx])
        if not any(cell.value for cell in row):
            continue
            
        total_rows += 1
        
        comp_name = get_row_val(row, indices["company_name"])
        raw_program_name = get_row_val(row, indices["program_name"])
        
        if not comp_name or not raw_program_name:
            invalid_rows_count += 1
            warnings.append({
                "row": r_idx,
                "message": f"Row skipped: Missing required parameters (Company Name or Program Name)."
            })
            continue

        valid_rows_count += 1
        
        # Parse fields
        c_category = get_row_val(row, indices["customer_category"]) or "Corporate"
        budget = parse_float(get_row_val(row, indices["budget"]))
        start_date = parse_date(get_row_val(row, indices["event_date_start"])) or parse_date(get_row_val(row, indices["start_date"]))
        end_date = parse_date(get_row_val(row, indices["event_date_end"])) or parse_date(get_row_val(row, indices["end_date"]))
        quote_no = get_row_val(row, indices["quotation_number"])
        
        p_status = map_program_status(get_row_val(row, indices["program_status"]))
        q_status = map_quotation_status(get_row_val(row, indices["quotation_status"]))
        pay_status = map_payment_status(get_row_val(row, indices["payment_status"]))
        proj_status = map_project_status(get_row_val(row, indices["project_status"]))
        raw_event_category = get_row_val(row, indices["event_category"])
        cancel_reason = get_row_val(row, indices["cancel_reason"])

        proj_title = build_project_title(comp_name, raw_event_category, raw_program_name)

        # 1. PO/PM initials lookup
        po_val = get_row_val(row, indices["po"])
        pm_val = get_row_val(row, indices["pm"])
        po_id = resolve_user_by_name_or_initials(db, po_val)
        pm_id = resolve_user_by_name_or_initials(db, pm_val)
        
        if po_val and not po_id:
            warnings.append({"row": r_idx, "message": f"Unmapped PO initials/name: '{po_val}'"})
        if pm_val and not pm_id:
            warnings.append({"row": r_idx, "message": f"Unmapped PM initials/name: '{pm_val}'"})

        # 2. Duplicate customer candidates
        duplicates = check_duplicate_customer(db, comp_name)
        dup_candidates = [d for d in duplicates if d.company_name.strip().lower() != comp_name.strip().lower()]
        if dup_candidates:
            warnings.append({
                "row": r_idx,
                "message": f"Duplicate customer candidate: '{comp_name}' matches existing customer '{dup_candidates[0].company_name}'"
            })

        # 3. Source type check
        partner_val = get_row_val(row, indices["partner"])
        sales_val = get_row_val(row, indices["sales"])
        combined_val = ((str(partner_val or "").strip()) + " " + (str(sales_val or "").strip())).strip().lower()
        is_unknown_source = False
        if combined_val:
            if not any(x in combined_val for x in ["hotel", "direct", "repeater", "partner", "instagram", "ig", "web", "website", "other"]):
                matched = False
                for ct in ["Hotel", "Direct", "Repeater", "Partner", "Instagram", "Web", "Other"]:
                    if ct.lower() == combined_val:
                        matched = True
                        break
                if not matched:
                    is_unknown_source = True
        if is_unknown_source:
            warnings.append({
                "row": r_idx,
                "message": f"Unknown source type for '{partner_val or sales_val}', falling back to 'Other'."
            })

        # 4. Project validation warnings simulation
        photo_val = get_row_val(row, indices["photo_links"])
        video_val = get_row_val(row, indices["video_links"])
        teaser_val = get_row_val(row, indices["teaser_links"])
        signed_val = get_row_val(row, indices["signed_document"])
        ig_val = get_row_val(row, indices["ig_links"])
        yt_val = get_row_val(row, indices["yt_links"])
        
        simulated_paid = 0.0
        if pay_status == "Paid":
            simulated_paid = float(budget or 0.0)
            
        docs_list = []
        if photo_val or video_val or teaser_val or signed_val or ig_val or yt_val:
            docs_list = [Document(file_path="simulated")]
            
        temp_proj = TempProject(
            budget=budget,
            event_date_start=start_date,
            event_date_end=end_date,
            quotation_status=q_status,
            program_status=p_status,
            payment_status=pay_status,
            project_status=proj_status,
            cancel_reason=cancel_reason,
            simulated_paid=simulated_paid
        )
        temp_proj.documents = docs_list
        
        # Simulate Project Instruments for validation warnings checking in preview (Sprint 7)
        sim_instruments = []
        for chk in ["cl", "ros", "ck", "pnl", "pf", "matrix"]:
            chk_val = get_row_val(row, indices[chk])
            inst_status, inst_doc_url, inst_notes, warning_msg = map_excel_instrument_val(chk_val)
            if warning_msg:
                warnings.append({
                    "row": r_idx,
                    "message": f"{warning_msg} for instrument {chk.upper()} (Set to 'In Progress' with notes)."
                })
            sim_instruments.append(ProjectInstrument(
                instrument_type=chk.upper(),
                status=inst_status,
                document_url=inst_doc_url,
                notes=inst_notes
            ))
        temp_proj.instruments = sim_instruments
        
        proj_warns = check_project_validation_warnings(temp_proj)
        for pw in proj_warns:
            warnings.append({
                "row": r_idx,
                "message": pw
            })

        preview_rows.append({
            "row_index": r_idx,
            "company_name": comp_name,
            "customer_category": c_category,
            "contact_name": get_row_val(row, indices["contact_name"]) or "Staff Contact",
            "phone": get_row_val(row, indices["phone"]),
            "email": get_row_val(row, indices["email"]),
            "project_title": proj_title,
            "budget": budget,
            "start_date": str(start_date) if start_date else None,
            "end_date": str(end_date) if end_date else None,
            "quotation_number": quote_no,
            "project_status": proj_status,
            "program_status": p_status,
            "payment_status": pay_status,
            "quotation_status": q_status
        })

        new_customers.add(comp_name)
        new_projects.add(f"{proj_title}-{start_date}")

    return {
        "total_rows": total_rows,
        "valid_rows_count": valid_rows_count,
        "invalid_rows_count": invalid_rows_count,
        "new_customers_count": len(new_customers),
        "new_projects_count": len(new_projects),
        "conflicts_count": conflicts_count,
        "warnings": warnings,
        "preview_rows": preview_rows
    }

def commit_excel_import(db: Session, file_bytes: bytes, user_id: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    
    sheet = None
    for s_name in wb.sheetnames:
        if clean_header(s_name) == "workflow" or s_name == "Workflow":
            sheet = wb[s_name]
            break
    if not sheet:
        sheet = wb.active

    indices, header_row_idx = find_column_indices(sheet)
    
    created_count = 0
    updated_count = 0
    errors = []
    
    # Trackers for quality reporting
    total_rows = 0
    successful_rows = 0
    skipped_rows = 0
    warning_count = 0
    error_count = 0
    warnings_per_row = {}
    unmapped_po_pm = []
    unknown_source_types = []
    duplicate_customers = []

    default_pm = db.query(User).filter(User.deleted_at == None).first()
    executor_id = user_id if user_id else (str(default_pm.id) if default_pm else None)

    for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
        row = list(sheet[r_idx])
        if not any(cell.value for cell in row):
            continue

        total_rows += 1
        comp_name = get_row_val(row, indices["company_name"])
        raw_program_name = get_row_val(row, indices["program_name"])

        if not comp_name or not raw_program_name:
            skipped_rows += 1
            error_count += 1
            errors.append(f"Row {r_idx} skipped: Missing required parameters (Company Name or Program Name).")
            continue

        row_warns = []
        try:
            # 1. Match/Create Customer & check duplicates
            customer = db.query(Customer).filter(
                Customer.company_name == comp_name,
                Customer.deleted_at == None
            ).first()

            if not customer:
                customer = Customer(
                    company_name=comp_name,
                    normalized_name=normalize_customer_name(comp_name),
                    category=get_row_val(row, indices["customer_category"]) or "Corporate",
                    address=get_row_val(row, indices["venue"]) or "Address Linked"
                )
                db.add(customer)
                db.commit()
                db.refresh(customer)
                created_count += 1
                
                log_activity(
                    db,
                    user_id=executor_id,
                    action="customer_created",
                    entity_type="customer",
                    entity_id=customer.id,
                    details={"company_name": customer.company_name, "category": customer.category, "source": "excel_sync"}
                )
            else:
                updated_count += 1

            # Duplicate Customer Candidate Check
            duplicates = check_duplicate_customer(db, comp_name)
            dup_candidates = [d for d in duplicates if d.company_name.strip().lower() != comp_name.strip().lower()]
            if dup_candidates:
                candidate = dup_candidates[0]
                msg = f"Duplicate customer candidate: '{comp_name}' matches existing customer '{candidate.company_name}'"
                row_warns.append(msg)
                duplicate_customers.append({
                    "row": r_idx,
                    "company_name": comp_name,
                    "candidate": candidate.company_name
                })

            # 2. Match/Create Contact Point
            c_name = get_row_val(row, indices["contact_name"])
            if c_name:
                contact = db.query(Contact).filter(
                    Contact.customer_id == customer.id,
                    Contact.name == c_name,
                    Contact.deleted_at == None
                ).first()
                if not contact:
                    contact = Contact(
                        customer_id=customer.id,
                        name=c_name,
                        email=get_row_val(row, indices["email"]),
                        phone=get_row_val(row, indices["phone"]),
                        position="Point of Contact"
                    )
                    db.add(contact)
                    db.commit()

            # 3. Resolve PO/PM from Initials or Names
            po_val = get_row_val(row, indices["po"])
            pm_val = get_row_val(row, indices["pm"])
            
            po_id = resolve_user_by_name_or_initials(db, po_val)
            pm_id = resolve_user_by_name_or_initials(db, pm_val)
            
            unresolved_notes = []
            if po_val and not po_id:
                unresolved_notes.append(f"Unresolved PO: {po_val}")
                row_warns.append(f"Unmapped PO initials/name: '{po_val}'")
                unmapped_po_pm.append({"row": r_idx, "field": "PO", "value": po_val})
            if pm_val and not pm_id:
                unresolved_notes.append(f"Unresolved PM: {pm_val}")
                row_warns.append(f"Unmapped PM initials/name: '{pm_val}'")
                unmapped_po_pm.append({"row": r_idx, "field": "PM", "value": pm_val})

            # 4. Resolve/Create Event Source (Controlled types normalization)
            partner_val = get_row_val(row, indices["partner"])
            sales_val = get_row_val(row, indices["sales"])
            event_source_id, source_warning = resolve_or_create_event_source(db, partner_val, sales_val)
            if source_warning:
                row_warns.append(source_warning)
                unknown_source_types.append({
                    "row": r_idx,
                    "value": str(partner_val or sales_val)
                })

            # 5. Match/Create Project
            proj_code = get_row_val(row, indices["project_code"])
            inquiry_date = parse_date(get_row_val(row, indices["inquiry_date"]))
            event_category = get_row_val(row, indices["event_category"])
            program_type = get_row_val(row, indices["program_type"])
            program_name = get_row_val(row, indices["program_name"]) or raw_program_name
            
            qty_raw = get_row_val(row, indices["quantity"])
            quantity = int(parse_float(qty_raw)) if qty_raw else None
            
            venue = get_row_val(row, indices["venue"])
            duration = get_row_val(row, indices["duration"])
            event_date_start = parse_date(get_row_val(row, indices["event_date_start"])) or parse_date(get_row_val(row, indices["start_date"]))
            event_date_end = parse_date(get_row_val(row, indices["event_date_end"])) or parse_date(get_row_val(row, indices["end_date"]))
            
            quotation_date = parse_date(get_row_val(row, indices["quotation_date"]))
            quotation_number = get_row_val(row, indices["quotation_number"])
            
            quotation_status = map_quotation_status(get_row_val(row, indices["quotation_status"]))
            program_status = map_program_status(get_row_val(row, indices["program_status"]))
            payment_status = map_payment_status(get_row_val(row, indices["payment_status"]))
            project_status = map_project_status(get_row_val(row, indices["project_status"]))
            
            budget = parse_float(get_row_val(row, indices["budget"]))
            cancel_reason = get_row_val(row, indices["cancel_reason"])
            mom_notes = get_row_val(row, indices["mom_notes"])
            
            general_notes = get_row_val(row, indices["general_notes"]) or ""
            if unresolved_notes:
                general_notes += "\n" + "; ".join(unresolved_notes)

            proj_title = build_project_title(comp_name, event_category, program_name)

            # Match criteria
            project = None
            if proj_code:
                project = db.query(Project).filter(
                    Project.project_code == str(proj_code).strip(),
                    Project.deleted_at == None
                ).first()
            
            if not project:
                project = db.query(Project).filter(
                    Project.title == proj_title,
                    Project.event_date_start == event_date_start,
                    Project.deleted_at == None
                ).first()

            if not project:
                project = Project(
                    title=proj_title,
                    status=program_status.lower(),
                    quotation_status=quotation_status,
                    start_date=event_date_start,
                    end_date=event_date_end,
                    budget=budget,
                    revenue=budget,
                    customer_id=customer.id,
                    created_by_id=uuid.UUID(executor_id) if executor_id else None,
                    assigned_to_id=pm_id or (uuid.UUID(executor_id) if executor_id else None),
                    
                    project_code=proj_code,
                    inquiry_date=inquiry_date,
                    event_category=event_category,
                    program_type=program_type,
                    program_name=program_name,
                    quantity=quantity,
                    venue=venue,
                    duration=duration,
                    event_date_start=event_date_start,
                    event_date_end=event_date_end,
                    quotation_date=quotation_date,
                    quotation_number=quotation_number,
                    program_status=program_status,
                    payment_status=payment_status,
                    project_status=project_status,
                    cancel_reason=cancel_reason,
                    mom_notes=mom_notes,
                    general_notes=general_notes,
                    
                    event_source_id=event_source_id,
                    program_owner_id=po_id,
                    program_manager_id=pm_id or (uuid.UUID(executor_id) if executor_id else None)
                )
                db.add(project)
                db.commit()
                db.refresh(project)
                created_count += 1
                
                log_activity(
                    db,
                    user_id=executor_id,
                    action="project_created",
                    entity_type="project",
                    entity_id=project.id,
                    details={"title": project.title, "status": project.program_status, "source": "excel_sync"}
                )
            else:
                project.title = proj_title
                project.start_date = event_date_start or project.start_date
                project.end_date = event_date_end or project.end_date
                project.budget = budget or project.budget
                project.revenue = budget or project.revenue
                project.status = program_status.lower()
                
                project.project_code = proj_code or project.project_code
                project.inquiry_date = inquiry_date or project.inquiry_date
                project.event_category = event_category or project.event_category
                project.program_type = program_type or project.program_type
                project.program_name = program_name or project.program_name
                project.quantity = quantity or project.quantity
                project.venue = venue or project.venue
                project.duration = duration or project.duration
                project.event_date_start = event_date_start or project.event_date_start
                project.event_date_end = event_date_end or project.event_date_end
                project.quotation_date = quotation_date or project.quotation_date
                project.quotation_number = quotation_number or project.quotation_number
                project.quotation_status = quotation_status or project.quotation_status
                project.program_status = program_status or project.program_status
                project.payment_status = payment_status or project.payment_status
                project.project_status = project_status or project.project_status
                project.cancel_reason = cancel_reason or project.cancel_reason
                project.mom_notes = mom_notes or project.mom_notes
                project.general_notes = general_notes or project.general_notes
                
                if event_source_id:
                    project.event_source_id = event_source_id
                if po_id:
                    project.program_owner_id = po_id
                if pm_id:
                    project.program_manager_id = pm_id
                    project.assigned_to_id = pm_id
                
                db.commit()
                updated_count += 1

            # 6. EventSchedule setup
            venue_name = venue or get_row_val(row, indices["venue"])
            if venue_name:
                sched = db.query(EventSchedule).filter(
                    EventSchedule.project_id == project.id,
                    EventSchedule.venue_name == venue_name,
                    EventSchedule.deleted_at == None
                ).first()
                if not sched:
                    sched = EventSchedule(
                        project_id=project.id,
                        venue_name=venue_name,
                        address=venue_name,
                        start_time=datetime.combine(event_date_start, datetime.min.time()) if event_date_start else datetime.now(),
                        end_time=datetime.combine(event_date_end, datetime.max.time()) if event_date_end else datetime.now(),
                        pic_id=pm_id or (uuid.UUID(executor_id) if executor_id else None),
                        rundown=[]
                    )
                    db.add(sched)
                    db.commit()

            # 7. Checklist Tasks syncing
            checklists = ["cl", "ros", "ck", "pnl", "pf", "matrix"]
            for chk in checklists:
                chk_val = get_row_val(row, indices[chk])
                chk_status = "done" if str(chk_val or "").strip().lower() in ["1", "yes", "check", "done", "x", "true"] else "todo"
                task_title = f"{chk.upper()} Operations Checklist"
                task = db.query(Task).filter(
                    Task.project_id == project.id,
                    Task.title == task_title,
                    Task.deleted_at == None
                ).first()
                
                if not task:
                    task = Task(
                        project_id=project.id,
                        title=task_title,
                        description=f"Compliance check for {chk.upper()}",
                        status=chk_status,
                        priority="medium",
                        due_date=datetime.combine(event_date_start, datetime.min.time()) if event_date_start else datetime.now(),
                        created_by_id=uuid.UUID(executor_id) if executor_id else None,
                        assigned_to_id=pm_id or (uuid.UUID(executor_id) if executor_id else None)
                    )
                    db.add(task)
                    db.commit()
                else:
                    if task.status != chk_status:
                        task.status = chk_status
                        db.commit()

            # 7.5. Project Instruments syncing (CL, ROS, CK, PNL, PF, MATRIX) (Sprint 7)
            checklists = ["cl", "ros", "ck", "pnl", "pf", "matrix"]
            for chk in checklists:
                chk_val = get_row_val(row, indices[chk])
                instrument_type = chk.upper()
                inst_status, inst_doc_url, inst_notes, warning_msg = map_excel_instrument_val(chk_val)
                
                if warning_msg:
                    row_warns.append(f"{warning_msg} for instrument {instrument_type}. Set to 'In Progress' with notes.")
                
                inst = db.query(ProjectInstrument).filter(
                    ProjectInstrument.project_id == project.id,
                    ProjectInstrument.instrument_type == instrument_type,
                    ProjectInstrument.deleted_at == None
                ).first()
                
                if not inst:
                    inst = ProjectInstrument(
                        project_id=project.id,
                        instrument_type=instrument_type,
                        status=inst_status,
                        title=instrument_type,
                        document_url=inst_doc_url,
                        notes=inst_notes,
                        updated_by_user_id=pm_id or (uuid.UUID(executor_id) if executor_id else None)
                    )
                    # For Sprint 7: Done status sets completed_date automatically
                    if inst_status == "Done":
                        inst.completed_date = date.today()
                    db.add(inst)
                    db.commit()
                else:
                    updated = False
                    if inst.status != inst_status:
                        inst.status = inst_status
                        if inst_status == "Done" and not inst.completed_date:
                            inst.completed_date = date.today()
                        # Prefer not to clear completed_date automatically in Sprint 7.
                        updated = True
                    if inst_doc_url and inst.document_url != inst_doc_url:
                        inst.document_url = inst_doc_url
                        updated = True
                    if inst_notes and inst.notes != inst_notes:
                        inst.notes = inst_notes
                        updated = True
                    
                    if updated:
                        inst.updated_by_user_id = pm_id or (uuid.UUID(executor_id) if executor_id else None)
                        inst.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
                        db.commit()

            # 8. Document Asset links syncing (improved standard types mapping)
            docs_mapping = {
                "photo_links": ("PHOTO", "Operational Photo Link"),
                "video_links": ("VIDEO", "Operational Video Link"),
                "teaser_links": ("TEASER", "Operational Teaser Link"),
                "ig_links": ("INSTAGRAM", "Instagram Link"),
                "yt_links": ("YOUTUBE", "YouTube Link"),
            }
            has_doc = False
            for doc_key, (doc_type, default_title) in docs_mapping.items():
                link_val = get_row_val(row, indices[doc_key])
                if link_val and str(link_val).startswith("http"):
                    has_doc = True
                    doc = db.query(Document).filter(
                        Document.project_id == project.id,
                        Document.document_type == doc_type,
                        Document.deleted_at == None
                    ).first()
                    if not doc:
                        doc = Document(
                            project_id=project.id,
                            title=default_title,
                            file_path=str(link_val).strip(),
                            file_type="link",
                            storage_type="google_drive",
                            uploaded_by_id=uuid.UUID(executor_id) if executor_id else None,
                            document_type=doc_type,
                            url=str(link_val).strip()
                        )
                        db.add(doc)
                        db.commit()

            # 9. Invoice & Payment syncing
            if quotation_number:
                invoice = db.query(Invoice).filter(
                    Invoice.invoice_number == str(quotation_number).strip(),
                    Invoice.deleted_at == None
                ).first()
                
                inv_status = "unpaid"
                if payment_status == "Paid":
                    inv_status = "paid"
                elif payment_status == "Partial Paid":
                    inv_status = "partial"
                
                pay_schedule = get_row_val(row, indices["payment_schedule"]) or "Excel Import Billing"
                
                if not invoice:
                    invoice = Invoice(
                        project_id=project.id,
                        invoice_number=str(quotation_number).strip(),
                        amount=budget,
                        issue_date=event_date_start or date.today(),
                        due_date=event_date_end or event_date_start or date.today(),
                        status=inv_status,
                        notes=f"Payment Schedule: {pay_schedule}"
                    )
                    db.add(invoice)
                    db.commit()
                    db.refresh(invoice)
                else:
                    invoice.amount = budget or invoice.amount
                    invoice.status = inv_status
                    invoice.notes = f"Payment Schedule: {pay_schedule}"
                    db.commit()
                
                if invoice.status == "paid":
                    payment = db.query(Payment).filter(
                        Payment.invoice_id == invoice.id,
                        Payment.status == "approved",
                        Payment.deleted_at == None
                    ).first()
                    
                    if not payment:
                        payment = Payment(
                            invoice_id=invoice.id,
                            amount=invoice.amount,
                            payment_date=event_date_start or date.today(),
                            reference_number=f"Excel Sync - {quotation_number}",
                            status="approved",
                            proof_url="https://drive.google.com/drive/folders/onespirit"
                        )
                        db.add(payment)
                        db.commit()

            # 10. Signed Document Sync
            signed_val = get_row_val(row, indices["signed_document"])
            if signed_val:
                has_doc = True
                doc_title = "Signed Operations Contract File"
                doc = db.query(Document).filter(
                    Document.project_id == project.id,
                    Document.document_type == "SIGNED_FILE",
                    Document.deleted_at == None
                ).first()
                if not doc:
                    doc = Document(
                        project_id=project.id,
                        title=doc_title,
                        file_path=str(signed_val).strip(),
                        file_type="pdf" if "pdf" in str(signed_val).lower() else "link",
                        storage_type="google_drive" if str(signed_val).startswith("http") else "local",
                        uploaded_by_id=uuid.UUID(executor_id) if executor_id else None,
                        notes="Seeded signed contract agreement from Excel import",
                        document_type="SIGNED_FILE",
                        url=str(signed_val).strip()
                    )
                    db.add(doc)
                    db.commit()

            # Project validation warnings check post-creation
            simulated_paid = float(budget or 0.0) if payment_status == "Paid" else 0.0
            temp_proj = TempProject(
                budget=budget,
                event_date_start=event_date_start,
                event_date_end=event_date_end,
                quotation_status=quotation_status,
                program_status=program_status,
                payment_status=payment_status,
                project_status=project_status,
                cancel_reason=cancel_reason,
                simulated_paid=simulated_paid
            )
            temp_proj.documents = [Document(file_path="simulated")] if has_doc else []
            proj_warns = check_project_validation_warnings(temp_proj)
            row_warns.extend(proj_warns)

            if row_warns:
                warning_count += len(row_warns)
                warnings_per_row[str(r_idx)] = row_warns

            successful_rows += 1

        except Exception as e:
            db.rollback()
            error_count += 1
            errors.append(f"Row {r_idx} import failed: {str(e)}")

    return {
        "success": len(errors) == 0,
        "created_count": created_count,
        "updated_count": updated_count,
        "errors": errors,
        "total_rows": total_rows,
        "successful_rows": successful_rows,
        "skipped_rows": skipped_rows,
        "warning_count": warning_count,
        "error_count": error_count,
        "warnings_per_row": warnings_per_row,
        "unmapped_po_pm": unmapped_po_pm,
        "unknown_source_types": unknown_source_types,
        "duplicate_customers": duplicate_customers
    }
